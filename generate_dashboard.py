#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
投资仪表板数据整合与网页生成脚本

用法:
  python generate_dashboard.py          # 从源文件读取全部数据，生成 data.xlsx 和 index.html
  python generate_dashboard.py --from-excel   # 只从 data.xlsx 读取，快速生成 index.html
"""

import openpyxl
from openpyxl import Workbook
import html
import json
import os
import re
import sys
from datetime import datetime, date

# ==================== 配置路径 ====================
HOME_ASSET_FILE = r'C:\Users\Zhiyong\Nutstore\1\research\投资文件夹\财务计算\家庭资产记录202603.xlsx'
BJEX_FILE = r'C:\Users\Zhiyong\Nutstore\1\research\投资文件夹\公众号写作素材\北交所研究\北交所2025-2026打新数据汇总（含趋势图）.xlsx'
HK_IPO_FILE = r'C:\Users\Zhiyong\Nutstore\1\research\投资文件夹\公众号写作素材\新股招股金额估计\港股新股总体情况表_202510至今.xlsx'
ML_FILE = r'C:\Users\Zhiyong\Nutstore\1\research\投资文件夹\公众号写作素材\可转债机器学习实盘\GBRT_机器学习选债仪表盘.html'

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(OUTPUT_DIR, 'data.xlsx')
HTML_FILE = os.path.join(OUTPUT_DIR, 'index.html')


def read_home_calendar():
    """读取家庭资产记录中的投资日历"""
    wb = openpyxl.load_workbook(HOME_ASSET_FILE, data_only=False)
    ws = wb['1-投资日历']
    
    dates_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    dates = []
    for i, val in enumerate(dates_row):
        if isinstance(val, datetime):
            dates.append((i, val))
    
    events = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name or not isinstance(name, str):
            continue
        
        event_dates = {}
        for col_idx, dt in dates:
            if col_idx < len(row) and row[col_idx]:
                val = str(row[col_idx]).strip()
                if val:
                    event_dates[dt.strftime('%Y-%m-%d')] = val
        
        if event_dates:
            category = '港股'
            if '北交所' in name:
                category = '北交所'
            elif 'SpaceX' in name or '币安' in name or 'TGE' in name:
                category = '其他'
            
            clean_name = name.replace('港股新股-', '').replace('北交所新股-', '').strip()
            events.append({
                'name': clean_name,
                'category': category,
                'dates': event_dates
            })
    
    return events


def read_bjex_data():
    """读取北交所打新历史数据"""
    wb = openpyxl.load_workbook(BJEX_FILE, data_only=True)
    ws = wb['北交所打新数据']
    
    headers = []
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:
            headers = [str(c).replace('\n', '') if c else '' for c in row]
        elif i >= 2 and row[0]:
            data.append(list(row))
    
    return headers, data


def read_hk_ipo_data():
    """读取港股新股总体情况表"""
    wb = openpyxl.load_workbook(HK_IPO_FILE, data_only=True)
    ws = wb['总体情况表']
    
    headers = []
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:
            headers = [str(c).replace('\n', '') if c else '' for c in row]
        elif i >= 2 and row[0]:
            data.append(list(row))
    
    return headers, data


def read_ml_dashboard():
    """读取机器学习仪表盘HTML中的表格数据和图表"""
    with open(ML_FILE, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Extract tables
    tables = re.findall(r'<table[^>]*>.*?</table>', content, re.DOTALL)
    result = {}
    
    table_names = ['current_holdings', 'last_holdings', 'backtest', 'rolling_validation']
    for idx, (name, table) in enumerate(zip(table_names, tables)):
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.DOTALL)
        table_data = []
        for row in rows:
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL)
            clean_cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
            table_data.append(clean_cells)
        result[name] = table_data
    
    # Extract chart images
    charts = []
    for m in re.finditer(r'<div class="section-title">((?:(?!</div>).)*?)</div>\s*<img[^>]*>', content, re.DOTALL):
        title_html = m.group(1)
        title = re.sub(r'<[^>]+>', '', title_html).strip()
        img_tag = m.group(0).split('<img')[1]
        img_tag = '<img' + img_tag
        title = html.unescape(title)
        charts.append({'title': title, 'img_tag': img_tag})
    
    result['charts'] = charts
    return result


def process_bjex_data(bjex_data):
    """处理北交所数据：涨跌幅和收益率乘以100"""
    for row in bjex_data:
        for i in (6, 8, 9):  # 首日收盘涨跌幅, 正股/碎股年化收益率 (用户已删除均价涨跌幅列)
            if row[i] is not None:
                try:
                    row[i] = float(row[i]) * 100
                except (ValueError, TypeError):
                    pass
    return bjex_data


def process_hk_data(hk_data):
    """处理港股数据：中签率除以100"""
    for row in hk_data:
        for i in (13, 16, 19):  # 甲尾/乙头/顶头中签率
            if row[i] is not None:
                try:
                    row[i] = float(row[i]) / 100
                except (ValueError, TypeError):
                    pass
    return hk_data


def create_data_excel(calendar_events, bjex_headers, bjex_data, 
                       hk_headers, hk_data, ml_data):
    """创建整合数据Excel文件（数据已预处理）"""
    wb = Workbook()
    
    # 1. 港股日历
    ws1 = wb.active
    ws1.title = '港股日历'
    ws1.append(['事件名称', '日期', '事件类型'])
    for event in calendar_events:
        if event['category'] in ('港股', '其他'):
            for dt_str, evt_type in sorted(event['dates'].items()):
                ws1.append([event['name'], dt_str, evt_type])
    
    # 2. 北交所日历
    ws2 = wb.create_sheet('北交所日历')
    ws2.append(['事件名称', '日期', '事件类型'])
    for event in calendar_events:
        if event['category'] == '北交所':
            for dt_str, evt_type in sorted(event['dates'].items()):
                ws2.append([event['name'], dt_str, evt_type])
    
    # 3. 北交所历史数据（数据已预处理）
    ws3 = wb.create_sheet('北交所历史数据')
    clean_headers = [h for h in bjex_headers if h][:11]
    ws3.append(clean_headers)
    for row in bjex_data:
        clean_row = [row[i] for i in range(min(len(clean_headers), len(row)))]
        ws3.append(clean_row)
    
    # 4. 港股历史数据（数据已预处理，包含25列）
    ws4 = wb.create_sheet('港股历史数据')
    clean_hk_headers = [h for h in hk_headers if h][:25]
    ws4.append(clean_hk_headers)
    for row in hk_data:
        clean_row = []
        for i in range(min(len(clean_hk_headers), len(row))):
            val = row[i]
            if isinstance(val, str) and val.startswith('='):
                clean_row.append(None)
            else:
                clean_row.append(val)
        ws4.append(clean_row)
    
    # 5-8. 可转债数据
    ws5 = wb.create_sheet('可转债当前持仓')
    for row in ml_data['current_holdings']:
        ws5.append(row)
    
    ws6 = wb.create_sheet('可转债上期持仓')
    for row in ml_data['last_holdings']:
        ws6.append(row)
    
    ws7 = wb.create_sheet('可转债回测表现')
    for row in ml_data['backtest']:
        ws7.append(row)
    
    ws8 = wb.create_sheet('可转债滚动验证')
    for row in ml_data['rolling_validation']:
        ws8.append(row)
    
    wb.save(DATA_FILE)
    print(f"数据已保存到: {DATA_FILE}")


def format_date(val):
    """统一日期格式：去掉时间部分，统一为 YYYY-MM-DD"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # 处理带时间的字符串，如 "2026-04-08 00:00:00"
    if ' ' in s:
        s = s.split(' ')[0]
    # 统一斜杠格式，如 "2025/9/26" → "2025-09-26"
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3:
            s = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return s


def read_data_from_excel():
    """从已生成的 data.xlsx 读取所有数据"""
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    
    # 港股日历
    ws = wb['港股日历']
    hk_calendar = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, dt, evt = row[0], row[1], row[2]
        if name not in hk_calendar:
            hk_calendar[name] = {'name': name, 'category': '港股', 'dates': {}}
        hk_calendar[name]['dates'][format_date(dt)] = evt
    
    # 北交所日历
    ws = wb['北交所日历']
    bjex_calendar = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, dt, evt = row[0], row[1], row[2]
        if name not in bjex_calendar:
            bjex_calendar[name] = {'name': name, 'category': '北交所', 'dates': {}}
        bjex_calendar[name]['dates'][format_date(dt)] = evt
    
    calendar_events = list(hk_calendar.values()) + list(bjex_calendar.values())
    
    # 北交所历史数据
    ws = wb['北交所历史数据']
    bjex_rows = list(ws.iter_rows(values_only=True))
    bjex_headers = [str(c) if c else '' for c in bjex_rows[0]]
    bjex_data = []
    for r in bjex_rows[1:]:
        row = list(r)
        # 格式化日期列（上市日期、中签公布日期）
        if len(row) > 2:
            row[2] = format_date(row[2])
        if len(row) > 9:
            row[9] = format_date(row[9])
        bjex_data.append(row)
    
    # 港股历史数据
    ws = wb['港股历史数据']
    hk_rows = list(ws.iter_rows(values_only=True))
    hk_headers = [str(c) if c else '' for c in hk_rows[0]]
    hk_data = []
    for r in hk_rows[1:]:
        row = list(r)
        # 格式化日期列（招股开始、截止申购、上市日期）
        for col_idx in (3, 4, 5):
            if len(row) > col_idx:
                row[col_idx] = format_date(row[col_idx])
        hk_data.append(row)
    
    # ML数据
    ml_data = {}
    for sheet_name, key in [
        ('可转债当前持仓', 'current_holdings'),
        ('可转债上期持仓', 'last_holdings'),
        ('可转债回测表现', 'backtest'),
        ('可转债滚动验证', 'rolling_validation')
    ]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        ml_data[key] = [[str(c) if c is not None else '' for c in r] for r in rows]
    
    ml_data['charts'] = []
    return calendar_events, bjex_headers, bjex_data, hk_headers, hk_data, ml_data


def generate_html(calendar_events, bjex_headers, bjex_data,
                  hk_headers, hk_data, ml_data):
    """生成投资仪表板HTML"""
    
    today_str = date.today().strftime('%Y-%m-%d')
    
    calendar_json = json.dumps(calendar_events, ensure_ascii=False, default=str)
    
    bjex_clean = []
    for row in bjex_data:
        bjex_clean.append([str(c) if c is not None else '' for c in row[:10]])
    bjex_json = json.dumps(bjex_clean, ensure_ascii=False)
    
    hk_clean = []
    for row in hk_data:
        clean_row = []
        for i in range(min(25, len(row))):
            val = row[i]
            if isinstance(val, str) and val.startswith('='):
                clean_row.append('')
            else:
                clean_row.append(str(val) if val is not None else '')
        hk_clean.append(clean_row)
    hk_json = json.dumps(hk_clean, ensure_ascii=False)
    
    ml_json = json.dumps({k: v for k, v in ml_data.items() if k != 'charts'}, ensure_ascii=False)
    
    # Build chart HTML sections
    cumulative_chart_html = ''
    other_charts_html = ''
    for chart in ml_data.get('charts', []):
        card_html = f'''<div class="card">
                <h2>{chart['title']}</h2>
                <div style="text-align:center;">
                    {chart['img_tag']}
                </div>
            </div>'''
        if '累计收益' in chart['title']:
            cumulative_chart_html = card_html
        else:
            other_charts_html += card_html
    
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>投资仪表板</title>
    <style>
        :root {{
            --bg: #0f172a;
            --card-bg: #1e293b;
            --card-bg-hover: #334155;
            --text: #f1f5f9;
            --text-secondary: #94a3b8;
            --accent: #38bdf8;
            --accent-green: #4ade80;
            --accent-red: #f87171;
            --accent-orange: #fb923c;
            --accent-purple: #a78bfa;
            --border: #334155;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: var(--bg);
            color: var(--text);
            line-height: 1.6;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        header {{
            text-align: center;
            padding: 30px 0;
            border-bottom: 1px solid var(--border);
            margin-bottom: 30px;
        }}
        header h1 {{ font-size: 2.2rem; color: var(--accent); }}
        header p {{ color: var(--text-secondary); margin-top: 8px; }}
        
        .nav-tabs {{
            display: flex;
            gap: 10px;
            margin-bottom: 25px;
            flex-wrap: wrap;
            justify-content: center;
        }}
        .nav-tab {{
            padding: 10px 24px;
            background: var(--card-bg);
            border: 1px solid var(--border);
            border-radius: 8px;
            cursor: pointer;
            color: var(--text-secondary);
            font-size: 0.95rem;
            transition: all 0.2s;
        }}
        .nav-tab:hover {{ background: var(--card-bg-hover); }}
        .nav-tab.active {{ background: var(--accent); color: #0f172a; font-weight: bold; }}
        
        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; }}
        
        .card {{
            background: var(--card-bg);
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 20px;
            border: 1px solid var(--border);
        }}
        .card h2 {{
            font-size: 1.3rem;
            margin-bottom: 15px;
            color: var(--accent);
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .card h3 {{
            font-size: 1.1rem;
            margin: 15px 0 10px;
            color: var(--text-secondary);
        }}
        
        .section-title {{
            font-size: 1.1rem;
            color: var(--text);
            margin: 20px 0 10px;
            padding-bottom: 8px;
            border-bottom: 1px solid var(--border);
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.88rem;
            margin-top: 10px;
        }}
        th, td {{
            padding: 10px 12px;
            text-align: left;
            border-bottom: 1px solid var(--border);
        }}
        th {{
            background: rgba(56, 189, 248, 0.1);
            color: var(--accent);
            font-weight: 600;
            position: sticky;
            top: 0;
        }}
        tr:hover {{ background: rgba(255,255,255,0.03); }}
        td {{ color: var(--text-secondary); }}
        td.positive {{ color: var(--accent-green); }}
        td.negative {{ color: var(--accent-red); }}
        
        .calendar-section {{ margin-bottom: 25px; }}
        .calendar-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
            gap: 15px;
            margin-top: 15px;
        }}
        .calendar-item {{
            background: rgba(255,255,255,0.03);
            border-radius: 10px;
            padding: 15px;
            border-left: 4px solid var(--accent);
            transition: transform 0.2s;
        }}
        .calendar-item:hover {{ transform: translateY(-2px); background: rgba(255,255,255,0.05); }}
        .calendar-item.bjex {{ border-left-color: var(--accent-green); }}
        .calendar-item.other {{ border-left-color: var(--accent-orange); }}
        .calendar-item.past {{ opacity: 0.5; }}
        .calendar-item.today {{ border: 1px solid var(--accent); box-shadow: 0 0 15px rgba(56,189,248,0.15); }}
        .calendar-item h4 {{
            font-size: 1rem;
            margin-bottom: 8px;
            color: var(--text);
        }}
        .calendar-item .tag {{
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 0.75rem;
            margin-bottom: 10px;
            background: rgba(56,189,248,0.15);
            color: var(--accent);
        }}
        .calendar-item.bjex .tag {{
            background: rgba(74,222,128,0.15);
            color: var(--accent-green);
        }}
        .calendar-item.other .tag {{
            background: rgba(251,146,60,0.15);
            color: var(--accent-orange);
        }}
        .calendar-dates {{
            font-size: 0.85rem;
        }}
        .calendar-dates .date-row {{
            display: flex;
            justify-content: space-between;
            padding: 4px 0;
            border-bottom: 1px solid rgba(255,255,255,0.05);
        }}
        .calendar-dates .date-row:last-child {{ border-bottom: none; }}
        .calendar-dates .date-row.today-row {{
            background: rgba(56,189,248,0.1);
            border-radius: 4px;
            padding: 4px 8px;
            margin: 2px -4px;
        }}
        .date-label {{ color: var(--text-secondary); }}
        .date-value {{ color: var(--text); font-weight: 500; }}
        .date-value.today-text {{ color: var(--accent); font-weight: bold; }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
            gap: 15px;
            margin: 15px 0;
        }}
        .stat-box {{
            background: rgba(255,255,255,0.03);
            border-radius: 10px;
            padding: 18px;
            text-align: center;
        }}
        .stat-box .value {{
            font-size: 1.6rem;
            font-weight: bold;
            color: var(--accent);
        }}
        .stat-box .value.green {{ color: var(--accent-green); }}
        .stat-box .value.red {{ color: var(--accent-red); }}
        .stat-box .label {{
            font-size: 0.85rem;
            color: var(--text-secondary);
            margin-top: 5px;
        }}
        
        .scroll-table {{
            max-height: 600px;
            overflow-y: auto;
            border-radius: 8px;
        }}
        
        .chart-img {{
            max-width: 100%;
            height: auto;
            border-radius: 8px;
        }}
        
        .footer {{
            text-align: center;
            padding: 30px;
            color: var(--text-secondary);
            font-size: 0.85rem;
            border-top: 1px solid var(--border);
            margin-top: 30px;
        }}
        
        @media (max-width: 768px) {{
            .container {{ padding: 10px; }}
            header h1 {{ font-size: 1.5rem; }}
            .nav-tab {{ padding: 8px 16px; font-size: 0.85rem; }}
            th, td {{ padding: 8px; font-size: 0.8rem; }}
            .calendar-grid {{ grid-template-columns: 1fr; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📊 投资仪表板</h1>
            <p>数据更新日期：{today_str}</p>
        </header>
        
        <div class="nav-tabs">
            <div class="nav-tab active" onclick="switchTab('calendar')">📅 打新日历</div>
            <div class="nav-tab" onclick="switchTab('bjex')">🏢 北交所数据</div>
            <div class="nav-tab" onclick="switchTab('hk')">🇭🇰 港股数据</div>
            <div class="nav-tab" onclick="switchTab('bond')">📈 可转债ML</div>
        </div>
        
        <!-- 打新日历 -->
        <div id="calendar" class="tab-content active">
            <div class="card">
                <h2>📅 近期打新日历</h2>
                <div id="calendar-upcoming"></div>
                <div id="calendar-past"></div>
            </div>
        </div>
        
        <!-- 北交所数据 -->
        <div id="bjex" class="tab-content">
            <div class="card">
                <h2>🏢 北交所打新历史数据</h2>
                <div class="scroll-table">
                    <table id="bjex-table">
                        <thead><tr></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <!-- 港股数据 -->
        <div id="hk" class="tab-content">
            <div class="card">
                <h2>🇭🇰 港股新股总体情况（最近50只）</h2>
                <div class="stats-grid" id="hk-stats"></div>
                <div class="scroll-table">
                    <table id="hk-table">
                        <thead><tr></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <!-- 可转债ML -->
        <div id="bond" class="tab-content">
            <div class="card">
                <h2>🔮 当前预测持仓</h2>
                <div class="scroll-table">
                    <table id="ml-current">
                        <thead><tr></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="card">
                <h2>📋 上期持仓回顾</h2>
                <div class="scroll-table">
                    <table id="ml-last">
                        <thead><tr></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            {cumulative_chart_html}
            <div class="card">
                <h2>📊 回测表现</h2>
                <div class="scroll-table">
                    <table id="ml-backtest">
                        <thead><tr></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="card">
                <h2>🔄 滚动验证结果</h2>
                <div class="scroll-table">
                    <table id="ml-validation">
                        <thead><tr></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            {other_charts_html}
        </div>
        
        <div class="footer">
            <p>投资有风险，入市需谨慎。本仪表板仅供个人研究使用。</p>
        </div>
    </div>
    
    <script>
        // ==================== 数据 ====================
        const calendarData = {calendar_json};
        const bjexData = {bjex_json};
        const hkData = {hk_json};
        const mlData = {ml_json};
        const today = new Date();
        today.setHours(0,0,0,0);
        
        // ==================== Tab切换 ====================
        function switchTab(tabId) {{
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.nav-tab').forEach(el => el.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            event.target.classList.add('active');
        }}
        
        // ==================== 渲染日历 ====================
        function renderCalendar() {{
            const upcomingContainer = document.getElementById('calendar-upcoming');
            const pastContainer = document.getElementById('calendar-past');
            
            const upcoming = [];
            const past = [];
            
            calendarData.forEach(item => {{
                const dates = Object.entries(item.dates).sort((a,b) => new Date(a[0]) - new Date(b[0]));
                const lastDate = new Date(dates[dates.length-1][0]);
                const hasToday = dates.some(([d]) => {{
                    const dt = new Date(d);
                    dt.setHours(0,0,0,0);
                    return dt.getTime() === today.getTime();
                }});
                
                const enriched = {{...item, dates: dates, hasToday, lastDate}};
                if (lastDate < today && !hasToday) {{
                    past.push(enriched);
                }} else {{
                    upcoming.push(enriched);
                }}
            }});
            
            upcoming.sort((a,b) => {{
                const aFirst = new Date(a.dates[0][0]);
                const bFirst = new Date(b.dates[0][0]);
                return aFirst - bFirst;
            }});
            
            past.sort((a,b) => b.lastDate - a.lastDate);
            
            function renderSection(title, items, container) {{
                if (items.length === 0) return;
                let html = `<div class="calendar-section"><div class="section-title">${{title}}</div><div class="calendar-grid">`;
                items.forEach(item => {{
                    const cls = item.category === '北交所' ? 'bjex' : (item.category === '其他' ? 'other' : '');
                    const isPast = item.lastDate < today && !item.hasToday ? 'past' : '';
                    const isToday = item.hasToday ? 'today' : '';
                    html += `<div class="calendar-item ${{cls}} ${{isPast}} ${{isToday}}">
                        <span class="tag">${{item.category}}</span>
                        <h4>${{item.name}}</h4>
                        <div class="calendar-dates">`;
                    item.dates.forEach(([dt, evt]) => {{
                        const dtObj = new Date(dt);
                        dtObj.setHours(0,0,0,0);
                        const isTodayRow = dtObj.getTime() === today.getTime();
                        const rowClass = isTodayRow ? 'today-row' : '';
                        const textClass = isTodayRow ? 'today-text' : '';
                        html += `<div class="date-row ${{rowClass}}">
                            <span class="date-label">${{dt}}</span>
                            <span class="date-value ${{textClass}}">${{evt}}</span>
                        </div>`;
                    }});
                    html += '</div></div>';
                }});
                html += '</div></div>';
                container.innerHTML = html;
            }}
            
            renderSection('📍 当前 / 即将开始', upcoming, upcomingContainer);
            renderSection('📌 已结束', past, pastContainer);
        }}
        
        // ==================== 渲染表格 ====================
        function renderTable(tableId, headers, rows, fmtConfig = {{}}) {{
            const table = document.getElementById(tableId);
            const thead = table.querySelector('thead tr');
            const tbody = table.querySelector('tbody');
            
            thead.innerHTML = headers.map(h => `<th>${{h}}</th>`).join('');
            
            tbody.innerHTML = rows.map(row => {{
                let cells = row.map((val, cidx) => {{
                    let s = val !== null && val !== undefined ? String(val) : '';
                    
                    // Apply formatting config
                    if (fmtConfig[cidx] && s !== '') {{
                        const cfg = fmtConfig[cidx];
                        const num = parseFloat(s);
                        if (!isNaN(num)) {{
                            if (cfg.type === 'pct') {{
                                s = (num * 100).toFixed(cfg.digits || 2) + '%';
                            }} else if (cfg.type === 'pct_display') {{
                                s = num.toFixed(cfg.digits || 2) + '%';
                            }} else if (cfg.type === 'num') {{
                                s = num.toFixed(cfg.digits || 2);
                            }}
                        }}
                    }}
                    
                    let cls = '';
                    if (s.includes('%') && !isNaN(parseFloat(s))) {{
                        const num = parseFloat(s);
                        cls = num > 0 ? 'positive' : (num < 0 ? 'negative' : '');
                    }}
                    return `<td class="${{cls}}">${{s}}</td>`;
                }}).join('');
                return `<tr>${{cells}}</tr>`;
            }}).join('');
        }}
        
        // ==================== 渲染统计 ====================
        function renderBjexStats() {{
            const total = bjexData.length;
            if (total === 0) return;
            const avgClose = bjexData.reduce((sum, r) => sum + (parseFloat(r[6])||0), 0) / total;
            const avgAvg = bjexData.reduce((sum, r) => sum + (parseFloat(r[7])||0), 0) / total;
            const positive = bjexData.filter(r => (parseFloat(r[6])||0) > 0).length;
            
            document.getElementById('bjex-stats').innerHTML = `
                <div class="stat-box"><div class="value">${{total}}</div><div class="label">总新股数</div></div>
                <div class="stat-box"><div class="value green">${{avgClose.toFixed(2)}}%</div><div class="label">平均收盘涨幅</div></div>
                <div class="stat-box"><div class="value green">${{avgAvg.toFixed(2)}}%</div><div class="label">平均均价涨幅</div></div>
                <div class="stat-box"><div class="value">${{positive}}</div><div class="label">上涨数量</div></div>
            `;
        }}
        
        function renderHkStats() {{
            const total = hkData.length;
            if (total === 0) return;
            const avgFirst = hkData.reduce((sum, r) => {{
                const v = parseFloat(r[22]);
                return sum + (isNaN(v) ? 0 : v * 100);
            }}, 0) / total;
            const positive = hkData.filter(r => (parseFloat(r[22])||0) > 0).length;
            
            document.getElementById('hk-stats').innerHTML = `
                <div class="stat-box"><div class="value">${{total}}</div><div class="label">总新股数</div></div>
                <div class="stat-box"><div class="value ${{avgFirst>=0?'green':'red'}}">${{avgFirst.toFixed(2)}}%</div><div class="label">平均首日涨幅</div></div>
                <div class="stat-box"><div class="value">${{positive}}</div><div class="label">首日上涨数</div></div>
            `;
        }}
        
        // ==================== 初始化 ====================
        document.addEventListener('DOMContentLoaded', () => {{
            renderCalendar();
            
            // 北交所：data.xlsx中已处理为百分比数值（如246.14）
            const bjexHeaders = ['股票代码', '股票名称', '上市日期', '发行价格(元)', '正股门槛(万元)', '碎股门槛(万元)', '首日收盘涨跌幅(%)', '正股年化收益率(%)', '碎股年化收益率(%)', '中签公布日期'];
            const bjexFmt = {{6: {{type:'pct_display', digits:2}}, 7: {{type:'pct_display', digits:2}}, 8: {{type:'pct_display', digits:2}}}};
            renderTable('bjex-table', bjexHeaders, bjexData, bjexFmt);
            
            // 港股：data.xlsx中涨幅/回报率已存储为小数增长率，中签率已÷100
            const hkHeaders = ['序号', '股票代码', '公司名称', '招股开始', '截止申购', '上市日期', '发售价(港元)', '每手股数', '公开发售股数', '认购倍数', '有效申请数', '总申购金额(亿港元)', '甲尾档位(股)', '甲尾中签率', '甲尾本金(万港元)', '乙头档位(股)', '乙头中签率', '乙头本金(万港元)', '顶头档位(股)', '顶头中签率', '顶头本金(万港元)', '富途暗盘涨幅', '首日涨幅', '甲尾回报率', '乙头回报率'];
            const hkFmt = {{
                13: {{type:'num', digits:4}},  // 甲尾中签率（已÷100）
                16: {{type:'num', digits:4}},  // 乙头中签率
                19: {{type:'num', digits:4}},  // 顶头中签率
                21: {{type:'pct', digits:2}},  // 暗盘涨幅（小数→%）
                22: {{type:'pct', digits:2}},  // 首日涨幅
                23: {{type:'pct', digits:2}},  // 甲尾回报率
                24: {{type:'pct', digits:2}}   // 乙头回报率
            }};
            renderTable('hk-table', hkHeaders.slice(0, hkData[0]?.length || 25), hkData.slice(-50), hkFmt);
            renderHkStats();
            
            renderTable('ml-current', mlData.current_holdings[0], mlData.current_holdings.slice(1));
            renderTable('ml-last', mlData.last_holdings[0], mlData.last_holdings.slice(1));
            renderTable('ml-backtest', mlData.backtest[0], mlData.backtest.slice(1));
            renderTable('ml-validation', mlData.rolling_validation[0], mlData.rolling_validation.slice(1));
        }});
    </script>
</body>
</html>
'''
    
    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"网页已生成: {HTML_FILE}")


def main():
    from_excel = '--from-excel' in sys.argv
    
    print("=" * 50)
    print("投资仪表板数据整合与网页生成")
    print("=" * 50)
    
    if from_excel:
        print("\n模式: 从 data.xlsx 直接生成网页")
        if not os.path.exists(DATA_FILE):
            print(f"错误: 找不到 {DATA_FILE}")
            print("请先运行: python generate_dashboard.py")
            return
        
        print("\n[1/3] 从 data.xlsx 读取数据...")
        calendar_events, bjex_headers, bjex_data, hk_headers, hk_data, ml_data = read_data_from_excel()
        print(f"  -> 日历事件: {len(calendar_events)}")
        print(f"  -> 北交所数据: {len(bjex_data)} 条")
        print(f"  -> 港股数据: {len(hk_data)} 条")
        
        print("\n[2/3] 从 ML 仪表盘读取图表...")
        ml_full = read_ml_dashboard()
        ml_data['charts'] = ml_full.get('charts', [])
        print(f"  -> 读取到 {len(ml_data['charts'])} 张图表")
        
        print("\n[3/3] 生成HTML网页...")
        generate_html(calendar_events, bjex_headers, bjex_data,
                      hk_headers, hk_data, ml_data)
    else:
        print("\n模式: 从源文件读取全部数据")
        
        print("\n[1/5] 读取投资日历...")
        calendar_events = read_home_calendar()
        print(f"  -> 读取到 {len(calendar_events)} 个事件")
        
        print("\n[2/5] 读取北交所历史数据...")
        bjex_headers, bjex_data = read_bjex_data()
        print(f"  -> 读取到 {len(bjex_data)} 条记录")
        
        print("\n[3/5] 读取港股数据...")
        hk_headers, hk_data = read_hk_ipo_data()
        print(f"  -> 读取到 {len(hk_data)} 条新股记录")
        
        print("\n[4/5] 读取机器学习仪表盘...")
        ml_data = read_ml_dashboard()
        print(f"  -> 读取到 {len(ml_data)-1} 个表格, {len(ml_data.get('charts', []))} 张图表")
        
        print("\n[5/5] 处理数据并创建整合Excel...")
        bjex_data = process_bjex_data(bjex_data)
        hk_data = process_hk_data(hk_data)
        create_data_excel(calendar_events, bjex_headers, bjex_data,
                          hk_headers, hk_data, ml_data)
        
        print("\n[6/6] 生成HTML网页...")
        generate_html(calendar_events, bjex_headers, bjex_data,
                      hk_headers, hk_data, ml_data)
    
    print("\n" + "=" * 50)
    print("全部完成！请打开 index.html 查看仪表板。")
    print("=" * 50)


if __name__ == '__main__':
    main()
